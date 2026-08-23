"""Services du module 17 : credentials, webhooks et import/export."""

from __future__ import annotations

import csv
import hashlib
import hmac
import io
import ipaddress
import json
import secrets
import socket
import uuid
from collections import defaultdict, deque
from datetime import date, datetime, timedelta
from pathlib import Path
from threading import Lock
from typing import Any, Optional
from urllib.parse import urlparse

import httpx
from jose import JWTError, jwt
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.config import settings
from app.models.integration import (
    APICredential,
    DataTransferJob,
    IntegrationConnection,
    IntegrationSyncJob,
    OAuthClient,
    WebhookDelivery,
    WebhookEvent,
    WebhookSubscription,
)
from app.models.owner import Owner, OwnerType
from app.models.property import Property, PropertyStatus, PropertyType
from app.models.tenant import EmploymentStatus, Tenant, TenantStatus
from app.services.admin_security_service import decrypt_secret, encrypt_secret
from app.services.owner_service import generate_reference as owner_reference
from app.services.property_service import generate_reference as property_reference
from app.services.tenant_service import generate_reference as tenant_reference


API_SCOPES = {
    "properties:read", "properties:write",
    "tenants:read", "tenants:write",
    "owners:read", "owners:write",
    "webhooks:read", "webhooks:write",
    "imports:write", "exports:read", "*",
}
WEBHOOK_EVENTS = [
    "property.created", "property.updated", "tenant.created", "lease.signed",
    "payment.received", "payment.overdue", "ticket.created", "ticket.updated",
    "document.signed", "insurance.claim.created", "automation.completed",
]
PROVIDERS = {
    "sage": ("accounting", ["api_key"]),
    "quickbooks": ("accounting", ["client_id", "client_secret", "refresh_token"]),
    "xero": ("accounting", ["client_id", "client_secret", "refresh_token"]),
    "bank_sync": ("banking", ["client_id", "client_secret"]),
    "stripe": ("payment", ["api_key"]),
    "gocardless": ("payment", ["access_token"]),
    "paypal": ("payment", ["client_id", "client_secret"]),
    "docusign": ("signature", ["client_id", "client_secret"]),
    "yousign": ("signature", ["api_key"]),
    "sendgrid": ("email", ["api_key"]),
    "mailgun": ("email", ["api_key"]),
    "smtp": ("email", ["password"]),
    "twilio": ("sms", ["account_sid", "auth_token"]),
    "ovh": ("sms", ["application_key", "application_secret", "consumer_key"]),
    "aws_s3": ("storage", ["access_key_id", "secret_access_key"]),
    "google_cloud_storage": ("storage", ["service_account_json"]),
    "google_maps": ("mapping", ["api_key"]),
    "mapbox": ("mapping", ["access_token"]),
    "seloger": ("real_estate_portal", ["api_key"]),
    "leboncoin": ("real_estate_portal", ["api_key"]),
    "google_calendar": ("calendar", ["client_id", "client_secret", "refresh_token"]),
    "outlook": ("calendar", ["client_id", "client_secret", "refresh_token"]),
    "salesforce": ("crm", ["client_id", "client_secret", "refresh_token"]),
    "hubspot": ("crm", ["access_token"]),
    "power_bi": ("bi", ["client_id", "client_secret"]),
    "tableau": ("bi", ["personal_access_token"]),
}
CONNECTOR_CATALOG = {
    "triggers": [
        "property.created", "tenant.created", "lease.signed", "payment.received",
        "payment.overdue", "ticket.created", "document.signed",
    ],
    "actions": [
        "create_property", "create_tenant", "create_ticket", "create_contact",
        "send_notification", "upload_document", "trigger_workflow",
    ],
    "searches": ["find_property", "find_tenant", "find_lease", "find_ticket"],
}

_rate_buckets: dict[str, deque] = defaultdict(deque)
_rate_lock = Lock()


def utcnow() -> datetime:
    return datetime.utcnow()


def object_view(row, exclude: set[str] | None = None) -> dict:
    excluded = exclude or set()
    result = {}
    for column in row.__table__.columns:
        if column.name in excluded:
            continue
        value = getattr(row, column.name)
        if hasattr(value, "value"):
            value = value.value
        result[column.name] = value
    return result


def validate_scopes(scopes: list[str]) -> list[str]:
    values = sorted(set(scopes))
    unknown = set(values) - API_SCOPES
    if unknown:
        raise ValueError(f"Scopes inconnus : {', '.join(sorted(unknown))}")
    return values


def create_api_key(db: Session, data, actor: str) -> tuple[APICredential, str]:
    scopes = validate_scopes(data.scopes)
    token = f"gim_live_{secrets.token_urlsafe(32)}"
    digest = hashlib.sha256(token.encode()).hexdigest()
    row = APICredential(
        name=data.name, key_prefix=token[:18], key_hash=digest, scopes=scopes,
        rate_limit_per_minute=data.rate_limit_per_minute,
        expires_at=data.expires_at.replace(tzinfo=None) if data.expires_at else None,
        created_by=actor,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row, token


def create_oauth_client(db: Session, data, actor: str) -> tuple[OAuthClient, str]:
    scopes = validate_scopes(data.scopes)
    client_id = f"gim_{secrets.token_urlsafe(18)}"
    secret = secrets.token_urlsafe(40)
    row = OAuthClient(
        name=data.name, client_id=client_id,
        client_secret_hash=hashlib.sha256(secret.encode()).hexdigest(),
        scopes=scopes, rate_limit_per_minute=data.rate_limit_per_minute,
        token_lifetime_seconds=data.token_lifetime_seconds, created_by=actor,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row, secret


def issue_oauth_token(db: Session, client_id: str, client_secret: str, requested_scope: Optional[str]) -> dict:
    client = db.query(OAuthClient).filter(OAuthClient.client_id == client_id, OAuthClient.is_active == True).first()  # noqa: E712
    if not client or not hmac.compare_digest(client.client_secret_hash, hashlib.sha256(client_secret.encode()).hexdigest()):
        raise ValueError("Client OAuth2 invalide")
    requested = requested_scope.split() if requested_scope else list(client.scopes or [])
    if not set(requested).issubset(set(client.scopes or [])):
        raise PermissionError("Un scope demandé n'est pas autorisé pour ce client")
    now = utcnow()
    expires = now + timedelta(seconds=client.token_lifetime_seconds)
    token = jwt.encode(
        {
            "sub": client.client_id, "type": "api_access", "scopes": requested,
            "iat": now, "exp": expires,
        },
        settings.SECRET_KEY,
        algorithm="HS256",
    )
    client.last_used_at = now
    db.commit()
    return {"access_token": token, "token_type": "bearer", "expires_in": client.token_lifetime_seconds, "scope": " ".join(requested)}


def authenticate_api(db: Session, api_key: Optional[str], bearer_token: Optional[str]) -> dict:
    now = utcnow()
    if api_key:
        digest = hashlib.sha256(api_key.encode()).hexdigest()
        row = db.query(APICredential).filter(APICredential.key_hash == digest, APICredential.is_active == True).first()  # noqa: E712
        if not row or (row.expires_at and row.expires_at.replace(tzinfo=None) <= now):
            raise ValueError("Clé API invalide ou expirée")
        row.last_used_at = now
        db.commit()
        return {"subject": f"api_key:{row.id}", "scopes": row.scopes or [], "rate_limit": row.rate_limit_per_minute}
    if bearer_token:
        try:
            payload = jwt.decode(bearer_token, settings.SECRET_KEY, algorithms=["HS256"])
        except JWTError as exc:
            raise ValueError("Jeton OAuth2 invalide") from exc
        if payload.get("type") != "api_access":
            raise ValueError("Jeton OAuth2 client_credentials requis")
        client = db.query(OAuthClient).filter(OAuthClient.client_id == payload.get("sub"), OAuthClient.is_active == True).first()  # noqa: E712
        if not client:
            raise ValueError("Client OAuth2 désactivé")
        client.last_used_at = now
        db.commit()
        return {"subject": f"oauth:{client.id}", "scopes": payload.get("scopes") or [], "rate_limit": client.rate_limit_per_minute}
    raise ValueError("Authentification X-API-Key ou OAuth2 requise")


def enforce_scope(principal: dict, required_scope: str) -> None:
    scopes = set(principal.get("scopes") or [])
    if "*" not in scopes and required_scope not in scopes:
        raise PermissionError(f"Scope {required_scope} requis")


def consume_rate_limit(principal: dict) -> tuple[int, int, int]:
    key = principal["subject"]
    limit = int(principal.get("rate_limit") or 100)
    now = datetime.now().timestamp()
    with _rate_lock:
        bucket = _rate_buckets[key]
        while bucket and now - bucket[0] >= 60:
            bucket.popleft()
        if len(bucket) >= limit:
            retry_after = max(1, int(60 - (now - bucket[0])))
            raise OverflowError(str(retry_after))
        bucket.append(now)
        remaining = max(0, limit - len(bucket))
        reset = int(bucket[0] + 60)
    return limit, remaining, reset


def provider_catalog() -> list[dict]:
    labels = {
        "accounting": "Comptabilité", "banking": "Banque", "payment": "Paiement",
        "signature": "Signature", "email": "Email", "sms": "SMS", "storage": "Stockage",
        "mapping": "Cartographie", "real_estate_portal": "Portails immobiliers",
        "calendar": "Calendrier", "crm": "CRM", "bi": "Business intelligence",
    }
    return [
        {
            "provider": provider, "category": category, "category_label": labels[category],
            "required_credentials": required, "supports_sync": True,
            "connection_test": "configuration" if provider != "smtp" else "configuration_and_optional_live",
        }
        for provider, (category, required) in PROVIDERS.items()
    ]


def connection_view(row: IntegrationConnection) -> dict:
    data = object_view(row, {"encrypted_credentials"})
    data["credentials_configured"] = bool(row.encrypted_credentials)
    data["credential_fields"] = row.credential_fields or []
    return data


def _credential_blob(credentials: dict) -> Optional[bytes]:
    return encrypt_secret(json.dumps(credentials, ensure_ascii=False, sort_keys=True)) if credentials else None


def _credentials(row: IntegrationConnection) -> dict:
    if not row.encrypted_credentials:
        return {}
    raw = decrypt_secret(row.encrypted_credentials)
    try:
        return json.loads(raw or "{}")
    except json.JSONDecodeError:
        return {}


def create_connection(db: Session, data, actor: str) -> IntegrationConnection:
    provider = data.provider.lower().strip()
    if provider not in PROVIDERS:
        raise ValueError("Fournisseur d'intégration inconnu")
    category, required = PROVIDERS[provider]
    if db.query(IntegrationConnection).filter(
        IntegrationConnection.provider == provider,
        IntegrationConnection.name == data.name,
    ).first():
        raise ValueError("Une connexion portant ce nom existe déjà pour ce fournisseur")
    credentials = {key: value for key, value in data.credentials.items() if value}
    missing = [field for field in required if not credentials.get(field)]
    row = IntegrationConnection(
        name=data.name, provider=provider, category=category, config=data.config,
        encrypted_credentials=_credential_blob(credentials), credential_fields=sorted(credentials),
        status="not_configured" if missing else "configured",
        last_error=f"Identifiants manquants : {', '.join(missing)}" if missing else None,
        is_active=data.is_active, created_by=actor,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


def update_connection(db: Session, row: IntegrationConnection, data) -> IntegrationConnection:
    payload = data.model_dump(exclude_unset=True)
    credentials = payload.pop("credentials", None)
    for key, value in payload.items():
        setattr(row, key, value)
    if credentials is not None:
        merged = _credentials(row)
        merged.update({key: value for key, value in credentials.items() if value})
        row.encrypted_credentials = _credential_blob(merged)
        row.credential_fields = sorted(merged)
    _, required = PROVIDERS[row.provider]
    missing = [field for field in required if not _credentials(row).get(field)]
    row.status = "not_configured" if missing else "configured"
    row.last_error = f"Identifiants manquants : {', '.join(missing)}" if missing else None
    db.commit()
    db.refresh(row)
    return row


def test_connection(db: Session, row: IntegrationConnection) -> dict:
    """Validation locale honnête ; aucun succès distant n'est simulé."""
    credentials = _credentials(row)
    _, required = PROVIDERS[row.provider]
    missing = [field for field in required if not credentials.get(field)]
    row.last_tested_at = utcnow()
    if missing:
        row.status = "not_configured"
        row.last_error = f"Identifiants manquants : {', '.join(missing)}"
        result = {"status": "not_configured", "configuration_valid": False, "missing_credentials": missing}
    else:
        row.status = "configured"
        row.last_error = None
        result = {
            "status": "configured",
            "configuration_valid": True,
            "live_test_performed": False,
            "message": "Configuration complète. Le test distant sera réalisé par l'adaptateur du fournisseur lors de la synchronisation.",
        }
    db.commit()
    return result


def create_sync_job(db: Session, connection: IntegrationConnection, data) -> IntegrationSyncJob:
    row = IntegrationSyncJob(
        connection_id=connection.id, direction=data.direction, resource=data.resource,
        cursor=data.cursor, status="requires_provider_worker",
        details={
            "options": data.options,
            "message": "Synchronisation journalisée ; un worker/adaptateur fournisseur doit être déployé pour appeler l'API distante.",
        },
    )
    if connection.status != "configured" or not connection.is_active:
        row.status = "rejected"
        row.error = "Connexion inactive ou incomplète"
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


# ---------------------------------------------------------------------------
# Webhooks signés
# ---------------------------------------------------------------------------
def _safe_webhook_url(url: str) -> None:
    parsed = urlparse(url)
    if parsed.scheme not in ({"https"} if settings.ENVIRONMENT == "production" else {"https", "http"}):
        raise ValueError("Une URL HTTPS est obligatoire en production")
    if not parsed.hostname or parsed.username or parsed.password:
        raise ValueError("URL webhook invalide")
    if settings.ENVIRONMENT != "production" and parsed.hostname in {"localhost", "127.0.0.1"}:
        return
    try:
        addresses = socket.getaddrinfo(parsed.hostname, parsed.port or 443, proto=socket.IPPROTO_TCP)
    except socket.gaierror as exc:
        raise ValueError("Le domaine webhook ne peut pas être résolu") from exc
    for address in addresses:
        ip = ipaddress.ip_address(address[4][0])
        if ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_reserved or ip.is_multicast:
            raise ValueError("Les adresses privées ou réservées sont interdites pour les webhooks")


def webhook_view(row: WebhookSubscription) -> dict:
    data = object_view(row, {"encrypted_secret"})
    data["secret_configured"] = bool(row.encrypted_secret)
    return data


def create_webhook(db: Session, data, actor: str) -> tuple[WebhookSubscription, Optional[str]]:
    target_url = str(data.target_url)
    _safe_webhook_url(target_url)
    unknown = [event for event in data.events if event not in WEBHOOK_EVENTS and event != "*"]
    if unknown:
        raise ValueError(f"Événements webhook inconnus : {', '.join(unknown)}")
    generated = None if data.secret else secrets.token_urlsafe(32)
    secret = data.secret or generated
    row = WebhookSubscription(
        name=data.name, target_url=target_url, events=sorted(set(data.events)),
        encrypted_secret=encrypt_secret(secret), is_active=data.is_active, created_by=actor,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row, generated


def event_matches(pattern: str, event_type: str) -> bool:
    return pattern == "*" or pattern == event_type or (pattern.endswith(".*") and event_type.startswith(pattern[:-1]))


def create_webhook_event(
    db: Session,
    event_type: str,
    data: dict,
    idempotency_key: Optional[str] = None,
    deliver_now: bool = False,
) -> WebhookEvent:
    if idempotency_key:
        existing = db.query(WebhookEvent).filter(WebhookEvent.idempotency_key == idempotency_key, WebhookEvent.event_type == event_type).first()
        if existing:
            return existing
    event = WebhookEvent(
        event_id=f"evt_{uuid.uuid4().hex}", event_type=event_type, data=data,
        idempotency_key=idempotency_key,
    )
    db.add(event)
    db.flush()
    subscriptions = db.query(WebhookSubscription).filter(WebhookSubscription.is_active == True).all()  # noqa: E712
    deliveries = []
    for subscription in subscriptions:
        if any(event_matches(pattern, event_type) for pattern in subscription.events or []):
            delivery = WebhookDelivery(subscription_id=subscription.id, event_id=event.id, attempt=1)
            db.add(delivery)
            deliveries.append(delivery)
    db.commit()
    if deliver_now:
        for delivery in deliveries:
            deliver_webhook(db, delivery)
    db.refresh(event)
    return event


def _event_payload(event: WebhookEvent) -> dict:
    return {
        "id": event.event_id,
        "type": event.event_type,
        "api_version": event.api_version,
        "occurred_at": event.occurred_at.isoformat() if event.occurred_at else utcnow().isoformat(),
        "data": event.data or {},
    }


def deliver_webhook(db: Session, delivery: WebhookDelivery) -> WebhookDelivery:
    subscription = db.query(WebhookSubscription).filter(WebhookSubscription.id == delivery.subscription_id).first()
    event = db.query(WebhookEvent).filter(WebhookEvent.id == delivery.event_id).first()
    if not subscription or not event:
        delivery.status, delivery.error = "failed", "Abonnement ou événement introuvable"
        db.commit()
        return delivery
    if not subscription.is_active:
        delivery.status, delivery.error = "cancelled", "Abonnement désactivé"
        db.commit()
        return delivery
    _safe_webhook_url(subscription.target_url)
    payload = _event_payload(event)
    body = json.dumps(payload, ensure_ascii=False, separators=(",", ":"), sort_keys=True).encode()
    timestamp = str(int(datetime.now().timestamp()))
    secret = decrypt_secret(subscription.encrypted_secret) or ""
    signature = hmac.new(secret.encode(), f"{timestamp}.".encode() + body, hashlib.sha256).hexdigest()
    delivery.signature = f"t={timestamp},v1={signature}"
    try:
        response = httpx.post(
            subscription.target_url,
            content=body,
            headers={
                "Content-Type": "application/json",
                "User-Agent": "GestImmo-Webhooks/1.0",
                "X-Gestimmo-Event": event.event_type,
                "X-Gestimmo-Delivery": str(delivery.id),
                "X-Gestimmo-Signature": delivery.signature,
            },
            timeout=8.0,
            follow_redirects=False,
        )
        delivery.response_status = response.status_code
        delivery.response_body = response.text[:2000]
        if 200 <= response.status_code < 300:
            delivery.status, delivery.delivered_at, delivery.error = "delivered", utcnow(), None
            subscription.failure_count = 0
        else:
            raise RuntimeError(f"HTTP {response.status_code}")
    except Exception as exc:
        delivery.status = "failed"
        delivery.error = str(exc)[:1000]
        delivery.next_retry_at = utcnow() + timedelta(minutes=min(60, 2 ** min(delivery.attempt, 5)))
        subscription.failure_count = (subscription.failure_count or 0) + 1
        if subscription.failure_count >= 10:
            subscription.is_active = False
            subscription.disabled_reason = "Désactivé après 10 échecs consécutifs"
    db.commit()
    db.refresh(delivery)
    return delivery


def retry_delivery(db: Session, row: WebhookDelivery) -> WebhookDelivery:
    latest_attempt = db.query(WebhookDelivery).filter(
        WebhookDelivery.subscription_id == row.subscription_id,
        WebhookDelivery.event_id == row.event_id,
    ).count()
    retry = WebhookDelivery(
        subscription_id=row.subscription_id, event_id=row.event_id,
        attempt=latest_attempt + 1,
    )
    db.add(retry)
    db.commit()
    db.refresh(retry)
    return deliver_webhook(db, retry)


# ---------------------------------------------------------------------------
# Import, migration et export massifs
# ---------------------------------------------------------------------------
ENTITY_FIELDS = {
    "properties": ["reference", "type", "status", "title", "address", "postal_code", "city", "country", "living_area", "rooms", "bedrooms", "rent_price", "charges", "sale_price"],
    "tenants": ["reference", "status", "first_name", "last_name", "email", "phone", "mobile", "address", "postal_code", "city", "employment_status", "monthly_net_income", "other_monthly_income"],
    "owners": ["reference", "owner_type", "first_name", "last_name", "company_name", "email", "phone", "mobile", "address", "postal_code", "city", "siret"],
}
REQUIRED_FIELDS = {
    "properties": {"type", "title", "address", "postal_code", "city"},
    "tenants": {"first_name", "last_name", "email"},
    "owners": set(),
}
ALIASES = {
    "titre": "title", "type_de_bien": "type", "adresse": "address", "code_postal": "postal_code",
    "ville": "city", "surface": "living_area", "pieces": "rooms", "pièces": "rooms",
    "loyer": "rent_price", "prix_de_vente": "sale_price", "charges": "charges",
    "prenom": "first_name", "prénom": "first_name", "nom": "last_name", "courriel": "email",
    "telephone": "phone", "téléphone": "phone", "revenu": "monthly_net_income",
    "societe": "company_name", "société": "company_name",
}


def _normalise_header(value: str) -> str:
    import unicodedata

    text = unicodedata.normalize("NFKD", str(value or "").strip().casefold())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return "_".join(filter(None, "".join(char if char.isalnum() else " " for char in text).split()))


def _read_tabular(path: Path, source_format: str) -> tuple[list[str], list[dict]]:
    if source_format == "csv":
        raw = path.read_bytes()
        text = raw.decode("utf-8-sig", errors="replace")
        try:
            dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        columns = list(reader.fieldnames or [])
        return columns, [dict(row) for row in reader]
    if source_format == "xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        try:
            columns = [str(value or "") for value in next(iterator)]
        except StopIteration:
            return [], []
        rows = [{columns[index]: value for index, value in enumerate(values)} for values in iterator]
        workbook.close()
        return columns, rows
    raise ValueError("Format d'import non pris en charge")


def _suggest_mapping(columns: list[str], entity_type: str) -> dict[str, str]:
    available = set(ENTITY_FIELDS[entity_type])
    mapping = {}
    for column in columns:
        normalized = _normalise_header(column)
        target = normalized if normalized in available else ALIASES.get(normalized)
        if target in available:
            mapping[column] = target
    return mapping


def analyse_import(db: Session, filename: str, content: bytes, entity_type: str, source_system: Optional[str], actor: str) -> DataTransferJob:
    if entity_type not in ENTITY_FIELDS:
        raise ValueError("Type d'entité non pris en charge")
    suffix = Path(filename).suffix.lower().lstrip(".")
    if suffix not in {"csv", "xlsx"}:
        raise ValueError("Seuls les fichiers CSV et XLSX sont acceptés")
    if len(content) > 20 * 1024 * 1024:
        raise ValueError("Fichier trop volumineux (20 Mo maximum)")
    folder = settings.private_upload_dir_path / "integrations"
    folder.mkdir(parents=True, exist_ok=True)
    ref = f"IMP-{uuid.uuid4().hex[:12].upper()}"
    path = folder / f"{ref}.{suffix}"
    path.write_bytes(content)
    columns, rows = _read_tabular(path, suffix)
    if not columns:
        path.unlink(missing_ok=True)
        raise ValueError("Le fichier ne contient aucun en-tête")
    mapping = _suggest_mapping(columns, entity_type)
    row = DataTransferJob(
        reference=ref, operation="migration" if source_system else "import",
        entity_type=entity_type, source_format=suffix, source_system=source_system,
        filename=filename, storage_path=str(path), mapping=mapping,
        status="analysed", columns=columns, preview=rows[:10], total_rows=len(rows),
        created_by=actor,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


def _parse_float(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    return float(str(value).replace(" ", "").replace(",", "."))


def _parse_int(value: Any) -> Optional[int]:
    parsed = _parse_float(value)
    return int(parsed) if parsed is not None else None


def _mapped_row(raw: dict, mapping: dict[str, str]) -> dict:
    # Le format principal est colonne source -> champ cible. L'inverse est aussi accepté.
    source_keys = set(raw)
    if not set(mapping).intersection(source_keys) and set(mapping.values()).intersection(source_keys):
        mapping = {source: target for target, source in mapping.items()}
    return {target: raw.get(source) for source, target in mapping.items() if target}


def _find_duplicate(db: Session, entity_type: str, data: dict):
    if data.get("reference"):
        model = {"properties": Property, "tenants": Tenant, "owners": Owner}[entity_type]
        row = db.query(model).filter(model.reference == data["reference"]).first()
        if row:
            return row
    if entity_type == "properties" and data.get("address") and data.get("postal_code"):
        return db.query(Property).filter(Property.address == data["address"], Property.postal_code == str(data["postal_code"])).first()
    if entity_type == "tenants" and data.get("email"):
        return db.query(Tenant).filter(Tenant.email == str(data["email"]).lower()).first()
    if entity_type == "owners" and data.get("email"):
        return db.query(Owner).filter(Owner.email == str(data["email"]).lower()).first()
    return None


def _clean_entity_data(entity_type: str, values: dict) -> dict:
    data = {key: value for key, value in values.items() if key in ENTITY_FIELDS[entity_type] and value not in (None, "")}
    for field in {"living_area", "rent_price", "charges", "sale_price", "monthly_net_income", "other_monthly_income"}:
        if field in data:
            data[field] = _parse_float(data[field])
    for field in {"rooms", "bedrooms"}:
        if field in data:
            data[field] = _parse_int(data[field])
    if "postal_code" in data:
        data["postal_code"] = str(data["postal_code"]).split(".")[0]
    if "email" in data:
        data["email"] = str(data["email"]).strip().lower()
    return data


def _coerce_enum_values(entity_type: str, values: dict) -> dict:
    data = dict(values)
    if entity_type == "properties":
        if data.get("type") is not None:
            data["type"] = PropertyType(data["type"])
        if data.get("status") is not None:
            data["status"] = PropertyStatus(data["status"])
    elif entity_type == "tenants":
        if data.get("status") is not None:
            data["status"] = TenantStatus(data["status"])
        if data.get("employment_status") is not None:
            data["employment_status"] = EmploymentStatus(data["employment_status"])
    elif data.get("owner_type") is not None:
        data["owner_type"] = OwnerType(data["owner_type"])
    return data


def _create_entity(db: Session, entity_type: str, values: dict):
    data = _coerce_enum_values(entity_type, values)
    data.pop("reference", None)
    if entity_type == "properties":
        data.setdefault("type", PropertyType.APARTMENT)
        data.setdefault("status", PropertyStatus.AVAILABLE)
        data.setdefault("country", "France")
        row = Property(reference=property_reference(), **data)
    elif entity_type == "tenants":
        data.setdefault("status", TenantStatus.ACTIVE)
        row = Tenant(reference=tenant_reference("TEN"), **data)
    else:
        data.setdefault("owner_type", OwnerType.INDIVIDUAL)
        row = Owner(reference=owner_reference(), **data)
    db.add(row)
    db.flush()
    return row


def execute_import(db: Session, job: DataTransferJob, mapping: dict, duplicate_strategy: str, dry_run: bool) -> dict:
    if job.operation not in {"import", "migration"} or not job.storage_path or not Path(job.storage_path).is_file():
        raise ValueError("Source d'import introuvable")
    effective_mapping = mapping or job.mapping or {}
    mapped_targets = set(effective_mapping.values())
    if not mapped_targets.intersection(ENTITY_FIELDS[job.entity_type]):
        mapped_targets = set(effective_mapping)
    missing = REQUIRED_FIELDS[job.entity_type] - mapped_targets
    if missing:
        raise ValueError(f"Mapping incomplet, champs requis : {', '.join(sorted(missing))}")
    _, rows = _read_tabular(Path(job.storage_path), job.source_format)
    stats = {"processed": 0, "created": 0, "updated": 0, "skipped": 0, "failed": 0}
    errors = []
    for line_number, raw in enumerate(rows, start=2):
        stats["processed"] += 1
        try:
            # Un SAVEPOINT par ligne empêche une contrainte SQL invalide de
            # rendre toute la session inutilisable et permet un rapport complet.
            with db.begin_nested():
                values = _clean_entity_data(job.entity_type, _mapped_row(raw, effective_mapping))
                missing_values = [field for field in REQUIRED_FIELDS[job.entity_type] if values.get(field) in (None, "")]
                if missing_values:
                    raise ValueError(f"Valeurs requises absentes : {', '.join(missing_values)}")
                duplicate = _find_duplicate(db, job.entity_type, values)
                if duplicate:
                    if duplicate_strategy == "error":
                        raise ValueError("Doublon détecté")
                    if duplicate_strategy == "skip":
                        stats["skipped"] += 1
                        continue
                    for field, value in _coerce_enum_values(job.entity_type, values).items():
                        if field != "reference":
                            setattr(duplicate, field, value)
                    stats["updated"] += 1
                else:
                    _create_entity(db, job.entity_type, values)
                    stats["created"] += 1
        except Exception as exc:
            stats["failed"] += 1
            if len(errors) < 500:
                errors.append({"row": line_number, "error": str(exc), "data": raw})
    if dry_run:
        db.rollback()
        return {"dry_run": True, **stats, "errors": errors}
    job.mapping = effective_mapping
    job.duplicate_strategy = duplicate_strategy
    job.status = "completed_with_errors" if stats["failed"] else "completed"
    job.processed_rows = stats["processed"]
    job.created_rows = stats["created"]
    job.updated_rows = stats["updated"]
    job.skipped_rows = stats["skipped"]
    job.failed_rows = stats["failed"]
    job.errors = errors
    job.completed_at = utcnow()
    db.commit()
    return {"job_id": job.id, "reference": job.reference, "dry_run": False, **stats, "errors": errors}


def _export_rows(db: Session, entity_type: str, filters: dict) -> list[dict]:
    model = {"properties": Property, "tenants": Tenant, "owners": Owner}[entity_type]
    query = db.query(model)
    if filters.get("city") and hasattr(model, "city"):
        query = query.filter(model.city.ilike(f"%{filters['city']}%"))
    if filters.get("status") and hasattr(model, "status"):
        query = query.filter(model.status == filters["status"])
    rows = []
    for item in query.limit(100_000).all():
        values = {}
        for field in ENTITY_FIELDS[entity_type]:
            value = getattr(item, field, None)
            if hasattr(value, "value"):
                value = value.value
            if isinstance(value, (datetime, date)):
                value = value.isoformat()
            values[field] = value
        rows.append(values)
    return rows


def export_data(db: Session, data, actor: str) -> tuple[DataTransferJob, bytes, str, str]:
    rows = _export_rows(db, data.entity_type, data.filters)
    fields = [field for field in (data.fields or ENTITY_FIELDS[data.entity_type]) if field in ENTITY_FIELDS[data.entity_type]]
    if not fields:
        raise ValueError("Aucun champ valide à exporter")
    reference = f"EXP-{uuid.uuid4().hex[:12].upper()}"
    folder = settings.private_upload_dir_path / "integrations" / "exports"
    folder.mkdir(parents=True, exist_ok=True)
    if data.output_format == "json":
        content = json.dumps([{field: row.get(field) for field in fields} for row in rows], ensure_ascii=False, indent=2).encode()
        media_type, suffix = "application/json", "json"
    elif data.output_format == "xlsx":
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = data.entity_type
        sheet.append(fields)
        for row in rows:
            sheet.append([row.get(field) for field in fields])
        stream = io.BytesIO()
        workbook.save(stream)
        content = stream.getvalue()
        media_type, suffix = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"
    else:
        stream = io.StringIO()
        writer = csv.DictWriter(stream, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
        content = ("\ufeff" + stream.getvalue()).encode("utf-8")
        media_type, suffix = "text/csv; charset=utf-8", "csv"
    output_path = folder / f"{reference}.{suffix}"
    output_path.write_bytes(content)
    job = DataTransferJob(
        reference=reference, operation="export", entity_type=data.entity_type,
        source_format=data.output_format, status="completed", mapping={"fields": fields},
        total_rows=len(rows), processed_rows=len(rows), output_path=str(output_path),
        created_by=actor, completed_at=utcnow(),
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    return job, content, media_type, output_path.name
