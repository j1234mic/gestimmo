# backend/app/routes/export.py

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
import csv as csv_module
import io
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm
import openpyxl

from app.database import get_db
from app.auth import require_read
from app.services.property_service import get_properties, get_property
from app.schemas.property import PropertyFilter
from app.models.property import PropertyEvaluation

router = APIRouter(prefix="/api/export", tags=["Export"])


@router.get("/properties/pdf")
def export_properties_pdf(
    search: Optional[str] = Query(None),
    type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user = Depends(require_read)
):
    """Exporter les biens en PDF."""
    
    filters = PropertyFilter(
        search=search,
        type=[type] if type else None,
        status=[status] if status else None,
        city=city
    )
    
    properties, total = get_properties(db, filters, 0, 1000)
    
    # Créer le PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#2563eb'),
        spaceAfter=20
    )
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=20
    )
    
    elements = []
    
    # Titre
    elements.append(Paragraph("Rapport - Biens Immobiliers", title_style))
    elements.append(Paragraph(
        f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} • {total} biens",
        subtitle_style
    ))
    
    # Tableau
    headers = ['Réf', 'Titre', 'Type', 'Ville', 'Surface', 'Prix', 'Statut']
    data = [headers]
    
    for prop in properties:
        price = ''
        if prop.get('rent_price'):
            price = f"{prop['rent_price']:,.0f} €/mois"
        elif prop.get('sale_price'):
            price = f"{prop['sale_price']:,.0f} €"
        
        data.append([
            prop.get('reference', '-'),
            prop.get('title', '-')[:40],
            prop.get('type', '-'),
            prop.get('city', '-'),
            f"{prop.get('living_area', '-')} m²",
            price,
            prop.get('status', '-')
        ])
    
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    # Ajuster les largeurs
    col_widths = [2.5*cm, 6*cm, 2*cm, 2.5*cm, 2*cm, 2.5*cm, 2*cm]
    table._argW = col_widths
    
    elements.append(table)
    elements.append(Spacer(1, 1*cm))
    elements.append(Paragraph(
        f"ImmoGest © {datetime.now().year} - Document généré automatiquement",
        subtitle_style
    ))
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=biens-{datetime.now().strftime('%Y%m%d')}.pdf"
        }
    )


@router.get("/properties/excel")
def export_properties_excel(
    search: Optional[str] = Query(None),
    type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user = Depends(require_read)
):
    """Exporter les biens en Excel."""
    
    filters = PropertyFilter(
        search=search,
        type=[type] if type else None,
        status=[status] if status else None,
        city=city
    )
    
    properties, total = get_properties(db, filters, 0, 1000)
    
    # Créer le classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Biens Immobiliers"
    
    # Styles
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
    header_fill = openpyxl.styles.PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    header_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    cell_alignment = openpyxl.styles.Alignment(vertical="center")
    
    # En-têtes
    headers = [
        'Référence', 'Titre', 'Type', 'Statut', 'Adresse', 'Code Postal', 'Ville',
        'Surface (m²)', 'Pièces', 'Chambres', 'SDB', 'Année',
        'Loyer (€)', 'Charges (€)', 'Prix vente (€)', 'DPE', 'Chauffage'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Données
    for row_idx, prop in enumerate(properties, 2):
        data = [
            prop.get('reference'),
            prop.get('title'),
            prop.get('type'),
            prop.get('status'),
            prop.get('address'),
            prop.get('postal_code'),
            prop.get('city'),
            prop.get('living_area'),
            prop.get('rooms'),
            prop.get('bedrooms'),
            prop.get('bathrooms'),
            prop.get('construction_year'),
            prop.get('rent_price'),
            prop.get('charges'),
            prop.get('sale_price'),
            prop.get('energy_class'),
            prop.get('heating_type'),
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = cell_alignment
    
    # Ajuster les largeurs
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Sauvegarder
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=biens-{datetime.now().strftime('%Y%m%d')}.xlsx"
        }
    )


@router.get("/properties/csv")
def export_properties_csv(
    search: Optional[str] = Query(None),
    type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user = Depends(require_read)
):
    """Exporter les biens en CSV."""
    filters = PropertyFilter(
        search=search,
        type=[type] if type else None,
        status=[status] if status else None,
        city=city,
    )

    properties, total = get_properties(db, filters, 0, 10000)

    buffer = io.StringIO()
    writer = csv_module.writer(buffer, delimiter=";")
    writer.writerow([
        "Référence", "Titre", "Type", "Statut", "Adresse", "Code Postal", "Ville",
        "Surface (m²)", "Pièces", "Chambres", "SDB", "Année", "Loyer (€)",
        "Charges (€)", "Prix vente (€)", "DPE", "GES", "Chauffage", "Tags",
    ])
    for prop in properties:
        writer.writerow([
            prop.get("reference"),
            prop.get("title"),
            prop.get("type"),
            prop.get("status"),
            prop.get("address"),
            prop.get("postal_code"),
            prop.get("city"),
            prop.get("living_area"),
            prop.get("rooms"),
            prop.get("bedrooms"),
            prop.get("bathrooms"),
            prop.get("construction_year"),
            prop.get("rent_price"),
            prop.get("charges"),
            prop.get("sale_price"),
            prop.get("energy_class"),
            prop.get("ges_class"),
            prop.get("heating_type"),
            ",".join(prop.get("tags") or []),
        ])

    data = buffer.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        io.BytesIO(data),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename=biens-{datetime.now().strftime('%Y%m%d')}.csv"
        },
    )


@router.get("/properties/{property_id}/evaluation-report")
def property_evaluation_report(
    property_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(require_read)
):
    """Générer le rapport d'évaluation PDF d'un bien."""
    property_obj = get_property(db, property_id)
    if not property_obj:
        raise HTTPException(status_code=404, detail="Bien non trouvé")

    evaluations = db.query(PropertyEvaluation).filter(
        PropertyEvaluation.property_id == property_id
    ).order_by(PropertyEvaluation.evaluation_date).all()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=2 * cm, leftMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "EvalTitle",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#2563eb"),
        spaceAfter=20,
    )
    subtitle_style = ParagraphStyle(
        "EvalSub",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.grey,
        spaceAfter=12,
    )

    elements = []
    elements.append(Paragraph(f"Rapport d'évaluation — {property_obj.title}", title_style))
    elements.append(Paragraph(
        f"Référence {property_obj.reference} • {property_obj.city}\n"
        f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
        subtitle_style,
    ))

    current = evaluations[-1].value if evaluations else None
    change = "—"
    if len(evaluations) >= 2:
        change = f"{current - evaluations[-2].value:+,.2f} €"
    info_rows = [
        ["Référence", property_obj.reference],
        ["Type", property_obj.type.value if property_obj.type else "-"],
        ["Statut", property_obj.status.value if property_obj.status else "-"],
        ["Surface habitable", f"{property_obj.living_area} m²" if property_obj.living_area else "-"],
        ["Surface totale", f"{property_obj.total_area} m²" if property_obj.total_area else "-"],
        ["Prix de vente", f"{property_obj.sale_price:,.2f} €" if property_obj.sale_price else "-"],
        ["Loyer mensuel", f"{property_obj.rent_price:,.2f} €" if property_obj.rent_price else "-"],
        ["Dernière évaluation", f"{current:,.2f} €" if current is not None else "-"],
        ["Évolution (2 dernières)", change],
    ]
    info = Table(info_rows, colWidths=[7 * cm, 11 * cm])
    info.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eff6ff")),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(info)
    elements.append(Spacer(1, 1 * cm))

    elements.append(Paragraph("Historique des évaluations", styles["Heading2"]))
    if not evaluations:
        elements.append(Paragraph("Aucune évaluation enregistrée.", styles["Normal"]))
    else:
        rows = [["Date", "Valeur", "Source", "Notes"]]
        for ev in evaluations:
            rows.append([
                ev.evaluation_date.strftime("%d/%m/%Y") if ev.evaluation_date else "-",
                f"{ev.value:,.2f} €",
                ev.source,
                (ev.notes or "")[:80],
            ])
        table = Table(rows, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        elements.append(table)

    elements.append(Spacer(1, 1 * cm))
    elements.append(Paragraph(
        "Document généré automatiquement — valeurs enregistrées dans GestImmo.",
        subtitle_style,
    ))
    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=rapport-evaluation-{property_obj.reference}.pdf"
        },
    )